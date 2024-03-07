# Quick function to multiply numeric values in an excel spreadsheet by 10
import argparse
import openpyxl

def multiplySpreadsheet(filename, factor):
    workbook = openpyxl.load_workbook(filename)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.value *= factor
    workbook.save(filename)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description = "Input the name of the file that you want to multiply by a value")
    parser.add_argument('filename', type=str, help="Name of the spreadsheet whose numbers you want to multiply")
    parser.add_argument('factor', type=float, help='The factor by which you want to multiply all the numbers on the spreadsheet')
    args = parser.parse_args()
    filename = args.filename
    factor = args.factor
    multiplySpreadsheet(filename, factor)